from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QLabel,
    QLineEdit,
    QDateEdit,
    QPushButton,
    QHBoxLayout,
    QGridLayout,
    QFileDialog,
    QTableView,
    QStyle,
    QTabWidget,
    QComboBox,
    QItemDelegate,
    QHeaderView,
    QSpacerItem,
    QSizePolicy,
    QCalendarWidget,
    QCheckBox,
    QMessageBox,
    QVBoxLayout,
    QAbstractItemView,
    QStyledItemDelegate
)
from PySide6.QtGui import QStandardItemModel, QStandardItem, QIcon
from PySide6.QtCore import Qt, QDate, QTimer
from pandasCGcalc import TransactionHistory, Portfolio
import sys
import pandas as pd
import datetime as dt
import os
import openpyxl as px
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

sys.argv += ['-platform', 'windows:darkmode=2']
expiredate = dt.date(2023, 12, 31)

class TransactionModel(QStandardItemModel):
    def data(self, index, role=Qt.DisplayRole): # type: ignore
        value = super().data(index, role)
        if index.column() in [4, 5, 8]:
            if role == Qt.DisplayRole: # type: ignore
                return '{:,.2f}'.format(float(value))
            elif role == Qt.TextAlignmentRole: # type: ignore
                return Qt.AlignRight | Qt.AlignCenter # type: ignore
        return value

    def flags(self, index):
        # This method is called with an index as an argument
        if index.column() in [8, 9]:  # I assume these are the columns you want to be read-only
            # You can use bitwise OR (|) to combine other flags as needed.
            return Qt.ItemIsEnabled | Qt.ItemIsSelectable # type: ignore
        else:
            return super().flags(index)  # For other columns, return the default flags

class CustomTableModel(QStandardItemModel):
    def data(self, index, role=Qt.DisplayRole): # type: ignore
        value = super().data(index, role)

        column_title = self.headerData(index.column(), Qt.Horizontal) # type: ignore Get the column title

        if column_title in ["Quantity", "Proceeds", "CostBase", "GrossValue", 'Value']: # Replace with your actual column titles
            if role == Qt.DisplayRole and value is not None: # type: ignore
                return '{:,.2f}'.format(float(value))
            elif role == Qt.TextAlignmentRole: # type: ignore
                return Qt.AlignRight | Qt.AlignCenter # type: ignore
        elif column_title in ["Discountable"]: # Replace with your actual column title
            if role == Qt.TextAlignmentRole: # type: ignore
                return Qt.AlignRight | Qt.AlignCenter # type: ignore
        return value
        
class TotalsModel(QStandardItemModel):
    def __init__(self, source_model):
        super().__init__(1, source_model.columnCount())
        self.source_model = source_model
        self.columns_to_sum = self.columns_to_sum = ['Value', 'Proceeds', 'CostBase', 'Quantity', 'GrossValue']
        self.source_model.rowsInserted.connect(self.update_totals)
        self.source_model.dataChanged.connect(self.update_totals)
        self.update_totals()
    
    def data(self, index, role=Qt.DisplayRole):  # type: ignore
        column_title = self.headerData(index.column(), Qt.Horizontal)  # type: ignore # Get the column title
        if column_title in self.columns_to_sum:
            if role == Qt.TextAlignmentRole:  # type: ignore
                return Qt.AlignRight | Qt.AlignVCenter  # type: ignore
        return super().data(index, role)

    def update_totals(self):
        self.setRowCount(1)
        self.setColumnCount(self.source_model.columnCount())
        for column in range(self.columnCount()):
            self.setItem(0, column, QStandardItem(''))
            column_title = self.source_model.headerData(column, Qt.Horizontal)  # type: ignore # Get the column title
            self.setHeaderData(column, Qt.Horizontal, column_title)  # type: ignore # Set the column header
            if column_title in self.columns_to_sum: 
                total = 0
                for row in range(self.source_model.rowCount()):
                    index = self.source_model.index(row, column)
                    value = index.data(Qt.DisplayRole)  # type: ignore
                    if value is not None and isinstance(value, str):
                        value = value.replace(',', '')  # Remove commas from the value
                        try:
                            total += float(value)
                        except ValueError:  # value can't be converted to float, skip it
                            continue
                self.setItem(0, column, QStandardItem(f'{float(total):,.2f}'))

class DeselectingLineEdit(QLineEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_table = None

    def setText(self, text):
        super().setText(text)
        QTimer.singleShot(0, self.deselect)

class TransactionHistoryEdit(QStyledItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.start_value = None
        self.parent_table = parent

    def setEditorData(self, editor, index):
        if self.start_value is not None:
            editor.setText(self.start_value) # type: ignore
            self.start_value = None
        else:
            super().setEditorData(editor, index)

    def createEditor(self, parent, option, index):
        editor = DeselectingLineEdit(parent)
        editor.parent_table = self.parent_table # type: ignore # Set the parent table of the editor

        # Check column title and set alignment
        column_title = self.parent_table.model().headerData(index.column(), Qt.Horizontal) # type: ignore
        if column_title in ['Quantity', 'Value']:
            editor.setAlignment(Qt.AlignmentFlag.AlignRight)
        return editor
    
    def setModelData(self, editor, model, index):
        text = editor.text() # type: ignore
        try:
            text = float(text)
            model.setData(index, text, Qt.EditRole) # type: ignore
        except:
            column_title = model.headerData(index.column(), Qt.Horizontal) # type: ignore
            if column_title == "Quantity":
                model.setData(index, 0.00, Qt.EditRole) # type: ignore
            elif column_title == "Value":
                model.setData(index, 0.00, Qt.EditRole) # type: ignore
            else:
                model.setData(index, text, Qt.EditRole) # type: ignore

class TransactionHistoryTable(QTableView):
    def __init__(self, mainWindow, parent=None):
        super().__init__(parent)
        self.mainWindow = mainWindow
        self.setItemDelegate(TransactionHistoryEdit(self))

class ComboBoxDelegate(QItemDelegate):
    def __init__(self, parent, items_list):
        super().__init__(parent)
        self.items_list = items_list

    def createEditor(self, parent, option, index):
        combo_box = QComboBox(parent)
        combo_box.addItems(self.items_list)
        return combo_box

    def setEditorData(self, editor, index):
        value = index.data(Qt.EditRole) # type: ignore
        editor.setCurrentText(value) # type: ignore

    def setModelData(self, editor, model, index):
        value = editor.currentText() # type: ignore
        model.setData(index, value, Qt.EditRole) # type: ignore

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect) # type: ignore

class CalendarDelegate(QItemDelegate):
    def __init__(self, parent=None):
        super().__init__(parent)

    def createEditor(self, parent, option, index):
        if index.column() == 0:  # Adjust the column index as per your requirement
            editor = QDateEdit(parent)
            editor.setCalendarPopup(True)
            editor.setDisplayFormat("yyyy-MM-dd")  
            return editor
        return super().createEditor(parent, option, index)

    def setEditorData(self, editor, index):
        if index.column() == 0:
            date_value = index.model().data(index, Qt.DisplayRole) # type: ignore
            try:
                if isinstance(date_value, str):
                    date_value = dt.date.fromisoformat(date_value)  # Convert string to dt.date
                qdate = QDate(date_value.year, date_value.month, date_value.day)  # Convert dt.date to QDate
                editor.setDate(qdate) # type: ignore
            except ValueError:
                # handle the exception for inappropriate date string
                editor.setDate(QDate.currentDate()) # type: ignore

    def setModelData(self, editor, model, index):
        if index.column() == 0:
            qdate = editor.date() # type: ignore
            dt_date = dt.date(qdate.year(), qdate.month(), qdate.day())  # Convert QDate to dt.date
            model.setData(index, dt_date.isoformat(), Qt.DisplayRole) # type: ignore

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect) # type: ignore

class CustomDateEdit(QDateEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Create a calendar widget
        self.calendar_widget = QCalendarWidget(self)
        self.calendar_widget.setWindowFlags(Qt.Popup) # type: ignore

        # Connect the calendar widget's selectionChanged signal to update the date edit
        self.calendar_widget.selectionChanged.connect(self.handleSelectionChanged)

        # Set the calendar widget as the calendar popup of the date edit
        self.setCalendarPopup(True)
        self.setCalendarWidget(self.calendar_widget)

    def handleSelectionChanged(self):
        # Get the selected date from the calendar widget
        selected_date = self.calendar_widget.selectedDate()

        # Set the selected date as the date of the date edit
        self.setDate(selected_date)

    def stepBy(self, steps):
        # Call the base class implementation to perform the default step behavior
        super().stepBy(steps)

        # Hide the calendar widget after stepping
        self.hideCalendar()

    def hideCalendar(self):
        # Hide the calendar widget
        self.calendar_widget.hide()

    def showCalendar(self):
        # Show the calendar widget
        self.calendar_widget.show()

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        
        self.transactionHistory = TransactionHistory()
        self.transactions = self.transactionHistory.transactions
        self.transactionHistoryModel = TransactionModel()
        self.transactionHistoryModel.setHorizontalHeaderLabels(self.transactions.columns.tolist())
        self.addRow()       
        
        self.taxDisplay = CustomTableModel()
        self.taxDisplay.setHorizontalHeaderLabels(['Date', 'AssetID', 'AssetType', 'TransactionType', 'Quantity', 'AcquisitionDate', 'Proceeds', 'CostBase', 'GrossValue', 'Discountable'])

        self.portfolioDisplay = CustomTableModel()
        self.portfolioDisplay.setHorizontalHeaderLabels(['AssetIdentifier', 'AssetType', 'OptionID', 'PurchaseDate', 'Quantity', 'Value', 'Discountable'])

        self.setWindowTitle("Capital Gains Calculator")
        self.setGeometry(100, 100, 1600, 1000)
        self.setMinimumWidth(1200)

        self.tabsWidget = QTabWidget(self)  

        # Create the first tab
        self.transactionHistoryTab = QWidget()
        self.tabsWidget.addTab(self.transactionHistoryTab, "Transaction History")

        # Create a layout for the first tab
        transactionHistoryLayout = QHBoxLayout(self.transactionHistoryTab)

        # Create a table view for first tab
        self.transactionHistoryView = TransactionHistoryTable(self)
        self.transactionHistoryView.setModel(self.transactionHistoryModel)
        self.transactionHistoryView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # type: ignore
        self.transactionHistoryView.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # type: ignore
        self.transactionHistoryView.setModel(self.transactionHistoryModel)
        self.transactionHistoryView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # type: ignore
        
        #Defining field edit items
        #0: Date, 1: AssetType, 2: AssetID, 3: TransactionType, 4: Quantity, 5: Value, 6: OptionID, 7: OptionSplitID, 8: GrossGain, 9: Discountable
        assetTypeList = ['Share', 'Option']
        transactionTypeList = ['Purchase', 'FIFO_Sale', 'LIFO_Sale', 'Split', 'Merge', 'Option_Sale', 'Exercise', 'Expire', 'Highest_Gain_Sale', 'Lowest_Gain_Sale']
        assetTypeComboBox = ComboBoxDelegate(self.transactionHistoryView, assetTypeList)
        transactionTypeComboBox = ComboBoxDelegate(self.transactionHistoryView, transactionTypeList)
        dateDelegate = CalendarDelegate(self.transactionHistoryView)
        self.transactionHistoryView.setItemDelegateForColumn(0, dateDelegate)
        self.transactionHistoryView.setItemDelegateForColumn(1, assetTypeComboBox)
        self.transactionHistoryView.setItemDelegateForColumn(3, transactionTypeComboBox)
        self.transactionHistoryView.verticalHeader().setSectionResizeMode(QHeaderView.Fixed) # type: ignore
        self.transactionHistoryView.verticalHeader().setFixedWidth(25)
        self.transactionHistoryView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        
        # Controls widet for first tab
        controls1 = QWidget()
        transactionHistoryControlsLayout = QGridLayout(controls1)
        transactionHistoryControlsLayout.setAlignment(Qt.AlignTop) # type: ignore
        controls1.setFixedWidth(400)

        # Add a file path label and line edit
        file_path_label = QLabel("File Path:")
        transactionHistoryControlsLayout.addWidget(file_path_label, 0, 0)

        self.filePathField = QLineEdit()
        transactionHistoryControlsLayout.addWidget(self.filePathField, 0, 1, 1, 8)

        # Add a button for opening a file dialog
        folder_button = QPushButton()
        pixmapi = QStyle.StandardPixmap.SP_DirIcon
        icon = self.style().standardIcon(pixmapi)
        folder_button.setIcon(icon)
        transactionHistoryControlsLayout.addWidget(folder_button, 0, 9)
        folder_button.clicked.connect(self.openFileDialog)

        # Add a button for reading transactions
        self.importTransactionsButton = QPushButton("Import Transactions")
        self.importTransactionsButton.clicked.connect(self.importTransactions)
        self.importTransactionsButton.setEnabled(False)
        transactionHistoryControlsLayout.addWidget(self.importTransactionsButton, 1, 0, 1, 10)
        self.filePathField.textChanged.connect(self.toggleReadTransactionsButton)
        
        # Add a button for saving changes within app
        self.saveChangesButton = QPushButton("Save Changes (Required before calculation)")
        self.saveChangesButton.setEnabled(False)
        self.saveChangesButton.clicked.connect(self.saveChanges)
        transactionHistoryControlsLayout.addWidget(self.saveChangesButton, 2, 0, 1, 10)
        self.transactionHistoryModel.itemChanged.connect(self.enableSaveButton)
        
        # Add a button for saving changes within app
        saveChangesToFileButton = QPushButton("Save Changes To File")
        saveChangesToFileButton.clicked.connect(self.saveChangesToFile)
        transactionHistoryControlsLayout.addWidget(saveChangesToFileButton, 3, 0, 1, 10)
        
        # Add new row button
        addRowButton = QPushButton("Add Row")
        addRowButton.clicked.connect(self.appendRow)
        transactionHistoryControlsLayout.addWidget(addRowButton, 4, 0, 1, 10)
        
        # Add remove row button
        removeRowButton = QPushButton("Remove Row")
        removeRowButton.clicked.connect(self.removeRow)
        transactionHistoryControlsLayout.addWidget(removeRowButton, 5, 0, 1, 10)
        
        # Add calculate button at bottom using spacer
        spacerToBottom = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding) # type: ignore
        transactionHistoryControlsLayout.addItem(spacerToBottom, 6, 0, 1, 10)
        self.calculate_button = QPushButton("Calculate")
        self.calculate_button.setFixedHeight(50)
        self.calculate_button.setEnabled(False)
        transactionHistoryControlsLayout.addWidget(self.calculate_button, 7, 0, 1, 10)

        # Add the table view and the controls to the layout of the first tab
        transactionHistoryLayout.addWidget(self.transactionHistoryView)
        transactionHistoryLayout.addWidget(controls1)

        # Create the second tab
        self.tab2 = QWidget()
        self.tabsWidget.addTab(self.tab2, "CGT Events")

        # Create a layout for the second tab
        cgtEventsLayout = QHBoxLayout(self.tab2)
        cgtEventsLayout.setContentsMargins(0, 0, 0, 0)

        # Create a table view for second tab
        self.cgtEventsView = QTableView()
        self.cgtEventsView.setModel(self.taxDisplay)
        self.cgtEventsView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.cgtEventsView.setModel(self.taxDisplay)
        self.cgtEventsView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # type: ignore
        self.cgtEventsView.verticalHeader().setSectionResizeMode(QHeaderView.Fixed) # type: ignore
        self.cgtEventsView.verticalHeader().setFixedWidth(25)
        self.cgtEventsView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.cgtEventsView.setEditTriggers(QAbstractItemView.NoEditTriggers) # type: ignore
        
        # Create total stable
        cgtEventsTotalsView = QTableView()
        self.cgtEventsTotalsModel = TotalsModel(self.taxDisplay)
        cgtEventsTotalsView.setModel(self.cgtEventsTotalsModel)
        cgtEventsTotalsView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        cgtEventsTotalsView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # type: ignore
        cgtEventsTotalsView.verticalHeader().setSectionResizeMode(QHeaderView.Fixed) # type: ignore
        cgtEventsTotalsView.verticalHeader().setFixedWidth(25)
        cgtEventsTotalsView.setFixedHeight(cgtEventsTotalsView.rowHeight(0) + cgtEventsTotalsView.horizontalHeader().height())
        cgtEventsTotalsView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        cgtEventsTotalsView.setEditTriggers(QAbstractItemView.NoEditTriggers) # type: ignore
        
        # Create a QVBoxLayout for stacking the tables
        cgtEventsStackedTablesLayout = QVBoxLayout()

        # Add both table views to the layout
        cgtEventsStackedTablesLayout.addWidget(self.cgtEventsView)
        cgtEventsStackedTablesLayout.addWidget(cgtEventsTotalsView)
        
        # Create a QWidget for the tables and set the layout
        cgtEventsStackedTablesWidget = QWidget()
        cgtEventsStackedTablesWidget.setLayout(cgtEventsStackedTablesLayout)
        
        # Connect calculate button to calculate function after second table has been instantiated
        self.calculate_button.clicked.connect(self.calculate)
        
        # Controls widet for second tab
        cgtEventsControls = QWidget()
        cgtEventsControlsLayout = QGridLayout(cgtEventsControls)
        cgtEventsControlsLayout.setAlignment(Qt.AlignTop) # type: ignore
        cgtEventsControls.setFixedWidth(400)
        
        # Apply date filter checkbox
        dateFilterLabel = QLabel('Apply date filter to transactions')
        self.dateFilterCheckbox = QCheckBox()
        cgtEventsControlsLayout.addWidget(dateFilterLabel, 0, 0, 1, 9)
        cgtEventsControlsLayout.addWidget(self.dateFilterCheckbox, 0, 9, 1, 1, Qt.AlignRight) # type: ignore
        
        # Add easy financial year selector to second controls layout
        self.financialYearSelector = QComboBox()
        self.financialYearSelector.addItems([str(year) for year in range(QDate.currentDate().year(), 2010, -1)])
        self.financialYearSelector.currentIndexChanged.connect(self.update_financial_year)
        financialYearSelectorLabel = QLabel("Select Financial Year:")
        cgtEventsControlsLayout.addWidget(financialYearSelectorLabel, 1, 0, 1, 1)
        cgtEventsControlsLayout.addWidget(self.financialYearSelector, 1, 1, 1, 9)

        # Tax period start label
        taxPeriodStartLabel = QLabel("Tax Period Start:")
        cgtEventsControlsLayout.addWidget(taxPeriodStartLabel, 2, 0)
        
        # Tax period start field
        self.taxPeriodStartField = CustomDateEdit()
        self.taxPeriodStartField.setMinimumDate(QDate())
        self.taxPeriodStartField.setMaximumDate(QDate()) 
        self.taxPeriodStartField.setDate(QDate(2022, 7, 1))
        cgtEventsControlsLayout.addWidget(self.taxPeriodStartField, 2, 1, 1, 9)

        # Tax period end label
        taxPeriodEndLabel = QLabel("Tax Period End:")
        cgtEventsControlsLayout.addWidget(taxPeriodEndLabel, 3, 0)

        # Tax period end field
        self.taxPeriodEndField = CustomDateEdit()
        self.taxPeriodEndField.setMinimumDate(QDate())
        self.taxPeriodEndField.setMaximumDate(QDate()) 
        self.taxPeriodEndField.setDate(QDate(2023, 6, 30))
        cgtEventsControlsLayout.addWidget(self.taxPeriodEndField, 3, 1, 1, 9)
        
        # Spacer before consolidation controls
        singleRowSpacer = QSpacerItem(20, 40)
        cgtEventsControlsLayout.addItem(singleRowSpacer, 4, 1, 1, 9)
        
        # Apply consolidation checkbox
        consolidationFilterLabel = QLabel('Apply consolidation to transactions')
        self.consolidationFilterCheckbox = QCheckBox()
        cgtEventsControlsLayout.addWidget(consolidationFilterLabel, 5, 0, 1, 9)
        cgtEventsControlsLayout.addWidget(self.consolidationFilterCheckbox, 5, 9, 1, 1, Qt.AlignRight) # type: ignore
        
        # Add consolidation selector
        self.consildationLevelSelector = QComboBox()
        self.consildationLevelSelector.addItems(['Date', 'Asset ID (Return-ready data)'])
        consolidationLevelSelectorLabel = QLabel("Consolidation level:")
        cgtEventsControlsLayout.addWidget(consolidationLevelSelectorLabel, 6, 0, 1, 1)
        cgtEventsControlsLayout.addWidget(self.consildationLevelSelector, 6, 1, 1, 9)
        
        # Spacer before export button
        cgtEventsControlsLayout.addItem(spacerToBottom, 7, 0, 1, 9)
        
        # Export button
        self.exportWorkpaperButton = QPushButton("Export Workpaper To Excel File")
        self.exportWorkpaperButton.setFixedHeight(50)
        cgtEventsControlsLayout.addWidget(self.exportWorkpaperButton, 8, 0, 1, 10)
        self.exportWorkpaperButton.clicked.connect(self.exportWorkpaper)
        
        # Add filter button at bottom using spacer
        filterButton = QPushButton("Filter")
        filterButton.setFixedHeight(50)
        cgtEventsControlsLayout.addWidget(filterButton, 9, 0, 1, 10)
        filterButton.clicked.connect(self.applyTaxFilter)

        # Add the table view and the controls to the layout of the first tab
        cgtEventsLayout.addWidget(cgtEventsStackedTablesWidget)
        cgtEventsLayout.addWidget(cgtEventsControls)
        
        # Create the third tab
        self.portfolioTab = QWidget()
        self.tabsWidget.addTab(self.portfolioTab, "Portfolio")

        # Create vertical box for stacked tables
        portfolioLayout = QVBoxLayout(self.portfolioTab)

        # Create a table view for third tab
        self.portfolioTableView = QTableView()
        self.portfolioTableView.setModel(self.portfolioDisplay) # TODO self.assetsDisplay
        self.portfolioTableView.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)  # type: ignore
        self.portfolioTableView.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)  # type: ignore
        self.portfolioTableView.setModel(self.portfolioDisplay) # TODO self.assetsDisplay
        self.portfolioTableView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # type: ignore
        self.portfolioTableView.verticalHeader().setSectionResizeMode(QHeaderView.Fixed) # type: ignore
        self.portfolioTableView.verticalHeader().setFixedWidth(25)
        self.portfolioTableView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.portfolioTableView.setEditTriggers(QAbstractItemView.NoEditTriggers) # type: ignore
        
        # Create totals row for portfolio tab
        portfolioTotalsView = QTableView()
        self.portfolioTotalsModel = TotalsModel(self.portfolioDisplay)
        portfolioTotalsView.setModel(self.portfolioTotalsModel)
        portfolioTotalsView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        portfolioTotalsView.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # type: ignore
        portfolioTotalsView.verticalHeader().setSectionResizeMode(QHeaderView.Fixed) # type: ignore
        portfolioTotalsView.verticalHeader().setFixedWidth(25)
        portfolioTotalsView.setFixedHeight(portfolioTotalsView.rowHeight(0) + portfolioTotalsView.horizontalHeader().height())
        portfolioTotalsView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        portfolioTotalsView.setEditTriggers(QAbstractItemView.NoEditTriggers) # type: ignore
        
        # Add the table view to the layout of the third tab
        portfolioLayout.addWidget(self.portfolioTableView)
        portfolioLayout.addWidget(portfolioTotalsView)
        

        self.setCentralWidget(self.tabsWidget)
        
        if dt.date.today() > expiredate:
            self.disableAll()
            expireMessage = QMessageBox()
            expireMessage.setIcon(QMessageBox.Warning) # type: ignore
            expireMessage.setText("Software expired")
            expireMessage.setWindowTitle("Alert")
            expireMessage.exec()
            

    def openFileDialog(self):
        file_dialog = QFileDialog(self)
        file_dialog.setFileMode(QFileDialog.ExistingFile)  # type: ignore
        file_path = file_dialog.getOpenFileName(self, "Select File")[0]
        if file_path:
            self.filePathField.setText(file_path)

    def importTransactions(self):
        self.transactionHistoryModel.setRowCount(0)
        self.transactionFilePath = self.filePathField.text()
        if not self.transactionFilePath:
            return
        self.transactionsFileName = os.path.basename(self.transactionFilePath)
        
        self.transactions = pd.read_csv(self.transactionFilePath)

        self.transactionHistory.readData(self.transactions)
        self.transactions = self.transactionHistory.transactions
        for i in self.transactions.index:
            for j in self.transactions.columns:
                item = QStandardItem(str(self.transactions.at[i, j]))
                self.transactionHistoryModel.setItem(i, self.transactions.columns.get_loc(j), item)

        self.transactionHistoryView.setModel(self.transactionHistoryModel)
    
    def saveChanges(self):
        df = pd.DataFrame()
        for i in range(self.transactionHistoryModel.rowCount()):
            for j in range(self.transactionHistoryModel.columnCount()):
                index = self.transactionHistoryModel.index(i, j)
                df.at[i, j] = self.transactionHistoryModel.data(index)
        df.columns = self.transactions.columns.tolist()
        
        self.transactionHistory.readData(df)
        self.transactions = self.transactionHistory.transactions
        self.saveChangesButton.setDisabled(True)
        self.calculate_button.setEnabled(True)
        
    def enableSaveButton(self):
        self.saveChangesButton.setDisabled(False)
        self.calculate_button.setDisabled(True)
    
    def saveChangesToFile(self):
        df = pd.DataFrame()
        for i in range(self.transactionHistoryModel.rowCount()):
            for j in range(self.transactionHistoryModel.columnCount()):
                index = self.transactionHistoryModel.index(i, j)
                df.at[i, j] = self.transactionHistoryModel.data(index)

        df.columns = self.transactions.columns.tolist()
        
        options = QFileDialog.Options() # type: ignore
        fileName, _ = QFileDialog.getSaveFileName(self,"Save As...", "","CSV Files (*.csv);;All Files (*)", options = options)
        if fileName:
            if '.csv' not in fileName:
                fileName += '.csv'
            df.to_csv(fileName, index=False)

    def addRow(self):
        #0: Date, 1: AssetType, 2: AssetID, 3: TransactionType, 4: Quantity, 5: Value, 6: OptionID, 7: OptionSplitID, 8: GrossGain, 9: Discountable
        row_count = self.transactionHistoryModel.rowCount()
        self.transactionHistoryModel.insertRow(row_count)
        self.transactionHistoryModel.setItem(row_count, 0, QStandardItem(dt.date.today().isoformat()))
        self.transactionHistoryModel.setItem(row_count, 1, QStandardItem('Share'))
        self.transactionHistoryModel.setItem(row_count, 3, QStandardItem('Purchase'))
        self.transactionHistoryModel.setItem(row_count, 4, QStandardItem('0.00'))
        self.transactionHistoryModel.setItem(row_count, 5, QStandardItem('0.00'))

    def removeRow(self):
        # Get current selection
        current_selection = self.transactionHistoryView.currentIndex()

        # Check if a row is selected
        if current_selection.row() >= 0:
            # Remove the row
            self.transactionHistoryModel.removeRow(current_selection.row())
        else:
            self.transactionHistoryModel.removeRow(self.transactionHistoryModel.rowCount() - 1)
        self.transactionHistoryView.selectRow(self.transactionHistoryModel.rowCount() - 1)

    def appendRow(self):
        # End editing the current item
        self.transactionHistoryView.closePersistentEditor(self.transactionHistoryView.currentIndex())
        # Append an empty row
        self.addRow()
        # Start editing the new item
        self.transactionHistoryView.edit(self.transactionHistoryModel.index(self.transactionHistoryModel.rowCount() - 1, 0))

    def keyPressEvent(self, event):
        if (event.key() == Qt.Key.Key_Return or event.key() == Qt.Key.Key_Enter) and self.transactionHistoryView.currentIndex().row() == self.transactionHistoryModel.rowCount() - 1:
            self.appendRow()
            self.transactionHistoryView.setCurrentIndex(self.transactionHistoryModel.index(self.transactionHistoryModel.rowCount() - 1, 0))

        elif event.key() == Qt.Key.Key_Tab and self.transactionHistoryView.currentIndex().row() == self.transactionHistoryModel.rowCount() - 1 and self.transactionHistoryView.currentIndex().column() == self.transactionHistoryModel.columnCount() - 1:
            self.appendRow()
            self.transactionHistoryView.setCurrentIndex(self.transactionHistoryModel.index(self.transactionHistoryModel.rowCount() - 1, 0))
        
        elif event.key() == Qt.Key.Key_Delete:
            self.removeRow()
            
    def toggleReadTransactionsButton(self):
        if self.filePathField.text():
            self.importTransactionsButton.setEnabled(True)
        else:
            self.importTransactionsButton.setEnabled(False)
    
    def calculate(self):
        self.portfolio = Portfolio() # Instantiate portfolio object
        self.portfolio.readTransactions(self.transactions) # Read transactions into portfolio based on transaction history
        self.taxTransactions = self.portfolio.taxableTransactions
        self.taxDisplay.setRowCount(0)
        for i in self.taxTransactions.index:
            for j in self.taxTransactions.columns:
                item = QStandardItem(str(self.taxTransactions.at[i, j]))
                self.taxDisplay.setItem(i, self.taxTransactions.columns.get_loc(j), item)

        self.cgtEventsView.setModel(self.taxDisplay) # Updates CGT event display
        self.calculate_button.setEnabled(False) # Disables calculate button once data has been calculated, until changes are saved ahain
        self.tabsWidget.setCurrentIndex(1) # Sets CGT event display as current tab view
        
        #Emit changes to Totals Model on CGT tab
        top_left_index = self.taxDisplay.index(0, 0)
        bottom_right_index = self.taxDisplay.index(self.taxDisplay.rowCount() - 1, self.taxDisplay.columnCount() - 1)
        self.taxDisplay.dataChanged.emit(top_left_index, bottom_right_index)
        
        self.assets = self.portfolio.consolidatePortfolio()
        self.portfolioDisplay.setHorizontalHeaderLabels(self.assets.columns.tolist())

        self.portfolioDisplay.setRowCount(0)
        for i in self.assets.index:
            for j in self.assets.columns:
                item = QStandardItem(str(self.assets.at[i, j]))
                self.portfolioDisplay.setItem(i, self.assets.columns.get_loc(j), item)

        self.portfolioTableView.setModel(self.portfolioDisplay)
        
    def update_financial_year(self, index):
        year = int(self.financialYearSelector.itemText(index))
        self.taxPeriodStartField.setDate(QDate(year-1, 7, 1))
        self.taxPeriodEndField.setDate(QDate(year, 6, 30))
    
    def applyTaxFilter(self):
        startDate = None
        endDate = None
        consolidationLevel = None
        if self.dateFilterCheckbox.isChecked():
            start = self.taxPeriodStartField.date()
            end = self.taxPeriodEndField.date()
            startDate = dt.date(start.year(), start.month(), start.day())
            endDate = dt.date(end.year(), end.month(), end.day())
        else:
            startDate = None
            endDate = None
        
        if self.consolidationFilterCheckbox.isChecked():
            consolidationLevel = self.consildationLevelSelector.currentIndex() + 1
        else:
            consolidationLevel = None
        
        if (not self.dateFilterCheckbox.isChecked()) and (not self.consolidationFilterCheckbox.isChecked()):
            self.taxDisplay.clear()
            self.taxDisplay.setHorizontalHeaderLabels(self.taxTransactions.columns.tolist())
            for i in self.taxTransactions.index:
                for j in self.taxTransactions.columns:
                    item = QStandardItem(str(self.taxTransactions.at[i, j]))
                    self.taxDisplay.setItem(i, self.taxTransactions.columns.get_loc(j), item)
            self.calculate()
            return
        
        self.filteredTaxTransactions = self.taxTransactions # instantiate filteredTaxTransactions before it hits other code
        
        self.filteredTaxTransactions.columns = self.taxTransactions.columns.tolist()
        self.filteredTaxTransactions = self.portfolio.filterTaxTransactions(self.taxTransactions, startDate, endDate, consolidationLevel)
        
        self.taxDisplay.clear()
        self.taxDisplay.setHorizontalHeaderLabels(self.filteredTaxTransactions.columns.tolist())
        for i in self.filteredTaxTransactions.index:
            for j in self.filteredTaxTransactions.columns:
                item = QStandardItem(str(self.filteredTaxTransactions.at[i, j]))
                self.taxDisplay.setItem(i, self.filteredTaxTransactions.columns.get_loc(j), item)

        self.cgtEventsView.setModel(self.taxDisplay)
        
        #Emit changes to the totals table on CGT Events tab
        top_left_index = self.taxDisplay.index(0, 0)
        bottom_right_index = self.taxDisplay.index(self.taxDisplay.rowCount() - 1, self.taxDisplay.columnCount() - 1)
        self.taxDisplay.dataChanged.emit(top_left_index, bottom_right_index)

    def exportWorkpaper(self):
        startDate = None
        endDate = None
        if self.dateFilterCheckbox.isChecked():
            start = self.taxPeriodStartField.date()
            end = self.taxPeriodEndField.date()
            startDate = dt.date(start.year(), start.month(), start.day())
            endDate = dt.date(end.year(), end.month(), end.day())
        else:
            startDate = None
            endDate = None
        
        tab1 = TransactionHistory.filterByDate(self.transactionHistory, self.transactions, startDate, endDate)
        tab2 = self.portfolio.filterTaxTransactions(self.taxTransactions, startDate, endDate)
        tab3 = self.portfolio.filterTaxTransactions(self.taxTransactions, startDate, endDate, 1)
        tab4 = self.portfolio.filterTaxTransactions(self.taxTransactions, startDate, endDate, 2)
        
        options = QFileDialog.Options() # type: ignore
        fileName, _ = QFileDialog.getSaveFileName(self,"Save As...", "","XLSX Files (*.xlsx);;All Files (*)", options = options)
        if fileName:
            if '.xlsx' not in fileName:
                fileName += '.xlsx'
        
        with pd.ExcelWriter(fileName) as writer:
            tab1.to_excel(writer, sheet_name = 'Transaction_Listing')
            tab2.to_excel(writer, sheet_name = 'CGT_Transactions')
            tab3.to_excel(writer, sheet_name = 'CGT_Consol_Date')
            tab4.to_excel(writer, sheet_name = 'CGT_Consol_Asset')
    
        wb = px.load_workbook(fileName)
        self.formatWorkpaper(wb)
        wb.save(fileName)
        
    def formatWorkpaper(self, workbook):
        for sheet in workbook.worksheets:
            headers = [cell.value for cell in sheet[1]]
            number_format_rules = {
                'Value': '#,##0.00',
                'Proceeds' : '#,##0.00',
                'CostBase' : '#,##0.00',
                'GrossValue' : '#,##0.00',
                'Date' : 'DD/MM/YYYY',
                'AcquisitionDate' : 'DD/MM/YYYY',
            }
            alignment_rules = {
                'Discountable' : Alignment(horizontal="left"),
            }
            for i, column_cells in enumerate(sheet.columns, start=1):
                column_letter = get_column_letter(i)
                for cell in column_cells:
                    # Apply number format if column is in number_format_rules
                    if headers[i-1] in number_format_rules:
                        cell.number_format = number_format_rules[headers[i-1]]
                    # Apply alignment if column is in alignment_rules
                    if headers[i-1] in alignment_rules:
                        cell.alignment = alignment_rules[headers[i-1]]
                
                max_length = 0
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 7)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    def disableAll(self):
        self.transactionHistoryTab.setDisabled(True)
        self.tab2.setDisabled(True)

app = QApplication(sys.argv)
app.setStyle('Fusion')
app.setWindowIcon(QIcon('C:/Users/mattt/Desktop/Programming/CostBaseApp/MoneySquare.png'))

window = MainWindow()
window.show()

app.exec()